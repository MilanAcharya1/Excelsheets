import openpyxl as xl

wb = xl.load_workbook('Invoice_1.xlsx')
first_sheet = wb['december22']
first_column_total = 0
second_column_total = 0
for row in range(1, first_sheet.max_row + 1):
    cell_product_quantity = first_sheet.cell(row, 4)
    cell_product_name = first_sheet.cell(row, 3)
    cell_product_price = first_sheet.cell(row, 5)
    try:
        if cell_product_quantity.value >= 1:                    #first side of products
            print(f'{cell_product_quantity.value} '
                  f'of {cell_product_name.value}'
                  f' for {cell_product_price.value}')
        each_total = cell_product_quantity.value * cell_product_price.value
        first_column_total = first_column_total + each_total
        print(each_total)
    except TypeError:
        pass
for row in range(1, first_sheet.max_row + 1):                   #2nd side of products
    cell_product_quantity = first_sheet.cell(row, 10)
    cell_product_name = first_sheet.cell(row, 9)
    cell_product_price = first_sheet.cell(row, 11)
    try:
        if cell_product_quantity.value >= 1:
            print(f'{cell_product_quantity.value} '
                  f'of {cell_product_name.value}'
                  f' for {cell_product_price.value}')
        each_total = cell_product_quantity.value * cell_product_price.value
        second_column_total = second_column_total + each_total
        print(each_total)
    except TypeError:
        pass
grand_total = (first_column_total + second_column_total)

print(f'grand total is: ', round(grand_total,2))